from flask import Flask, render_template, request, jsonify, Response
import json, os, math, copy, webbrowser, threading, csv, io, datetime, re

app = Flask(__name__)
BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
DATA_FILE      = os.path.join(BASE_DIR, 'portfolio.json')
SCENARIOS_DIR  = os.path.join(BASE_DIR, 'scenarios')

# Fallback constants (overridden by stored settings)
WACC      = 0.10
TAX_RATE  = 0.26
INFLATION = 0.03
HORIZON   = 10

RAMP = {
    'Fast':   [0.625, 1.0,  1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0],
    'Medium': [0.35,  0.75, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0],
    'Slow':   [0.25,  0.625,1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0],
}

WEIGHTS = {
    'FFCF': {'npv': 0.0,   'irr': 0.0,   'roce': 0.385, 'capex': 0.615},
    'EBIT': {'npv': 0.241, 'irr': 0.225, 'roce': 0.425, 'capex': 0.109},
}

DEFAULT_STAGES = {
    'Capital Investment': [
        {'name': 'Feasibility',          'duration_months': 3,  'color': '#6366f1'},
        {'name': 'Engineering & Design', 'duration_months': 9,  'color': '#3b82f6'},
        {'name': 'Procurement',          'duration_months': 6,  'color': '#f59e0b'},
        {'name': 'Construction',         'duration_months': 18, 'color': '#ef4444'},
        {'name': 'Commissioning',        'duration_months': 3,  'color': '#f97316'},
        {'name': 'Operations',           'duration_months': 81, 'color': '#16a34a'},
    ],
    'R&D / Innovation': [
        {'name': 'Research',      'duration_months': 12, 'color': '#6366f1'},
        {'name': 'Development',   'duration_months': 18, 'color': '#8b5cf6'},
        {'name': 'Testing',       'duration_months': 9,  'color': '#f59e0b'},
        {'name': 'Pilot',         'duration_months': 6,  'color': '#f97316'},
        {'name': 'Scale-up',      'duration_months': 12, 'color': '#10b981'},
        {'name': 'Steady State',  'duration_months': 63, 'color': '#16a34a'},
    ],
    'Digital / IT': [
        {'name': 'Requirements', 'duration_months': 3,  'color': '#6366f1'},
        {'name': 'Design',       'duration_months': 3,  'color': '#3b82f6'},
        {'name': 'Development',  'duration_months': 9,  'color': '#f59e0b'},
        {'name': 'Testing & QA', 'duration_months': 3,  'color': '#ef4444'},
        {'name': 'Deployment',   'duration_months': 3,  'color': '#f97316'},
        {'name': 'Support Ops',  'duration_months': 99, 'color': '#16a34a'},
    ],
    'Operations Improvement': [
        {'name': 'Assessment',      'duration_months': 2,  'color': '#6366f1'},
        {'name': 'Planning',        'duration_months': 3,  'color': '#3b82f6'},
        {'name': 'Implementation',  'duration_months': 9,  'color': '#f59e0b'},
        {'name': 'Stabilisation',   'duration_months': 6,  'color': '#f97316'},
        {'name': 'Steady State',    'duration_months': 100,'color': '#16a34a'},
    ],
    'Cost Reduction': [
        {'name': 'Analysis',        'duration_months': 3,  'color': '#6366f1'},
        {'name': 'Design',          'duration_months': 3,  'color': '#3b82f6'},
        {'name': 'Implementation',  'duration_months': 9,  'color': '#f59e0b'},
        {'name': 'Monitoring',      'duration_months': 105,'color': '#16a34a'},
    ],
    'Infrastructure': [
        {'name': 'Planning',         'duration_months': 6,  'color': '#6366f1'},
        {'name': 'Procurement',      'duration_months': 6,  'color': '#3b82f6'},
        {'name': 'Construction',     'duration_months': 24, 'color': '#ef4444'},
        {'name': 'Commissioning',    'duration_months': 3,  'color': '#f97316'},
        {'name': 'Operations',       'duration_months': 81, 'color': '#16a34a'},
    ],
    'ESG / Sustainability': [
        {'name': 'Assessment',      'duration_months': 3,  'color': '#16a34a'},
        {'name': 'Planning',        'duration_months': 3,  'color': '#3b82f6'},
        {'name': 'Implementation',  'duration_months': 18, 'color': '#f59e0b'},
        {'name': 'Certification',   'duration_months': 3,  'color': '#6366f1'},
        {'name': 'Steady State',    'duration_months': 93, 'color': '#10b981'},
    ],
    'M&A / Integration': [
        {'name': 'Due Diligence',    'duration_months': 3,  'color': '#6366f1'},
        {'name': 'Planning',         'duration_months': 3,  'color': '#3b82f6'},
        {'name': 'Integration',      'duration_months': 18, 'color': '#ef4444'},
        {'name': 'Optimisation',     'duration_months': 12, 'color': '#f97316'},
        {'name': 'Steady State',     'duration_months': 84, 'color': '#16a34a'},
    ],
    'Growth / Revenue': [
        {'name': 'Market Research',  'duration_months': 3,  'color': '#6366f1'},
        {'name': 'Go-to-Market',     'duration_months': 6,  'color': '#3b82f6'},
        {'name': 'Launch',           'duration_months': 6,  'color': '#f59e0b'},
        {'name': 'Scale',            'duration_months': 12, 'color': '#f97316'},
        {'name': 'Steady State',     'duration_months': 93, 'color': '#16a34a'},
    ],
}

def _stages(cat):
    return copy.deepcopy(DEFAULT_STAGES.get(cat, DEFAULT_STAGES['Capital Investment']))

DEFAULT_PROJECTS = [
    {'id':1,'name':'Project Aurora','description':'Utility-scale solar + battery storage complex, 100 GWh/yr output',
     'project_category':'Capital Investment','project_type':'New Build','status':'Approved','owner':'Alex Riordan',
     'budget_type':'CapEx','department':'Energy & Utilities','region':'Southwest','risk_level':'Medium','strategic_score':5,
     'budget':85.0,'capacity':100.0,'unit_label':'GWh','unit_price':80000.0,'unit_cost':22000.0,'efficiency':0.92,
     'ramp':'Slow','schedule_driven':False,'accelerate':False,'locked_year':None,'stages':_stages('Capital Investment')},
    {'id':2,'name':'Project Nexus','description':'Enterprise cloud ERP migration — 500 users, $3.5M annual productivity gain',
     'project_category':'Digital / IT','project_type':'Upgrade','status':'Active','owner':'Priya Mehta',
     'budget_type':'OpEx','department':'Information Technology','region':'Corporate HQ','risk_level':'Medium','strategic_score':4,
     'budget':12.0,'capacity':500.0,'unit_label':'users','unit_price':8500.0,'unit_cost':2200.0,'efficiency':0.90,
     'ramp':'Fast','schedule_driven':False,'accelerate':True,'locked_year':None,'stages':_stages('Digital / IT')},
    {'id':3,'name':'Project Titan','description':'Greenfield advanced manufacturing facility — 200 kT annual throughput',
     'project_category':'Capital Investment','project_type':'New Build','status':'Concept','owner':'Marcus Webb',
     'budget_type':'CapEx','department':'Manufacturing','region':'Midwest','risk_level':'High','strategic_score':5,
     'budget':95.0,'capacity':200.0,'unit_label':'kT','unit_price':175000.0,'unit_cost':55000.0,'efficiency':1.0,
     'ramp':'Slow','schedule_driven':False,'accelerate':False,'locked_year':None,'stages':_stages('Capital Investment')},
    {'id':4,'name':'Project Phoenix','description':'Brownfield plant modernisation — upgrade 150 kT facility, cut OPEX by 15%',
     'project_category':'Capital Investment','project_type':'Upgrade','status':'Approved','owner':'Sandra Cole',
     'budget_type':'CapEx','department':'Manufacturing','region':'Southeast','risk_level':'Low','strategic_score':4,
     'budget':40.0,'capacity':150.0,'unit_label':'kT','unit_price':165000.0,'unit_cost':48000.0,'efficiency':1.0,
     'ramp':'Fast','schedule_driven':False,'accelerate':True,'locked_year':None,'stages':_stages('Capital Investment')},
    {'id':5,'name':'Project Atlas','description':'End-to-end supply chain re-engineering — freight & inventory optimisation',
     'project_category':'Operations Improvement','project_type':'Improvement','status':'Active','owner':'James Okafor',
     'budget_type':'OpEx','department':'Supply Chain','region':'North America','risk_level':'Low','strategic_score':3,
     'budget':5.0,'capacity':8.0,'unit_label':'$M saved','unit_price':1000000.0,'unit_cost':80000.0,'efficiency':0.85,
     'ramp':'Fast','schedule_driven':False,'accelerate':True,'locked_year':None,'stages':_stages('Operations Improvement')},
    {'id':6,'name':'Project Horizon','description':'Next-generation product R&D platform — 4 new product lines over 5 years',
     'project_category':'R&D / Innovation','project_type':'R&D Program','status':'Concept','owner':'Dr. Lena Fischer',
     'budget_type':'Mixed','department':'R&D','region':'Northeast','risk_level':'High','strategic_score':5,
     'budget':28.0,'capacity':4.0,'unit_label':'products','unit_price':12000000.0,'unit_cost':3000000.0,'efficiency':0.70,
     'ramp':'Slow','schedule_driven':False,'accelerate':False,'locked_year':None,'stages':_stages('R&D / Innovation')},
    {'id':7,'name':'Project Catalyst','description':'Lean Six Sigma transformation — eliminate waste across 6 production lines',
     'project_category':'Cost Reduction','project_type':'Process Improvement','status':'Active','owner':'Tom Nakamura',
     'budget_type':'OpEx','department':'Operations','region':'Midwest','risk_level':'Low','strategic_score':3,
     'budget':3.5,'capacity':10.0,'unit_label':'$M saved','unit_price':1000000.0,'unit_cost':60000.0,'efficiency':0.90,
     'ramp':'Fast','schedule_driven':False,'accelerate':True,'locked_year':None,'stages':_stages('Cost Reduction')},
    {'id':8,'name':'Project Vanguard','description':'European market entry — launch premium product portfolio in 5 countries',
     'project_category':'Growth / Revenue','project_type':'Market Expansion','status':'Approved','owner':'Claire Dubois',
     'budget_type':'Mixed','department':'Commercial','region':'Europe','risk_level':'High','strategic_score':4,
     'budget':38.0,'capacity':80.0,'unit_label':'k customers','unit_price':1200.0,'unit_cost':350.0,'efficiency':0.80,
     'ramp':'Slow','schedule_driven':False,'accelerate':False,'locked_year':None,'stages':_stages('Growth / Revenue')},
    {'id':9,'name':'Project Echo','description':'Digital twin deployment — real-time plant simulation across 25 facilities',
     'project_category':'Digital / IT','project_type':'Platform','status':'Approved','owner':'Ravi Patel',
     'budget_type':'CapEx','department':'Digital & Technology','region':'Global','risk_level':'Medium','strategic_score':4,
     'budget':20.0,'capacity':25.0,'unit_label':'plants','unit_price':700000.0,'unit_cost':120000.0,'efficiency':0.95,
     'ramp':'Fast','schedule_driven':False,'accelerate':True,'locked_year':None,'stages':_stages('Digital / IT')},
    {'id':10,'name':'Project Zenith','description':'Phase-1 capacity expansion — brownfield add 180 kT at existing flagship site',
     'project_category':'Capital Investment','project_type':'Expansion','status':'Concept','owner':'Marcus Webb',
     'budget_type':'CapEx','department':'Manufacturing','region':'Northeast','risk_level':'Medium','strategic_score':4,
     'budget':60.0,'capacity':180.0,'unit_label':'kT','unit_price':160000.0,'unit_cost':50000.0,'efficiency':1.0,
     'ramp':'Slow','schedule_driven':False,'accelerate':False,'locked_year':None,'stages':_stages('Capital Investment')},
    {'id':11,'name':'Project Nova','description':'Greenfield regional distribution hub — 400 k sq ft automated facility',
     'project_category':'Infrastructure','project_type':'New Build','status':'Concept','owner':'Helen Marsh',
     'budget_type':'CapEx','department':'Logistics','region':'Central','risk_level':'Medium','strategic_score':3,
     'budget':48.0,'capacity':400.0,'unit_label':'k sq ft','unit_price':25.0,'unit_cost':6.0,'efficiency':0.88,
     'ramp':'Slow','schedule_driven':False,'accelerate':False,'locked_year':None,'stages':_stages('Infrastructure')},
    {'id':12,'name':'Project Apex','description':'Yield improvement program — reduce off-spec output from 8% to 2%',
     'project_category':'Cost Reduction','project_type':'Process Improvement','status':'Active','owner':'Tom Nakamura',
     'budget_type':'OpEx','department':'Manufacturing','region':'Southeast','risk_level':'Low','strategic_score':3,
     'budget':4.0,'capacity':120.0,'unit_label':'kT','unit_price':18000.0,'unit_cost':3500.0,'efficiency':1.0,
     'ramp':'Fast','schedule_driven':False,'accelerate':True,'locked_year':None,'stages':_stages('Cost Reduction')},
    {'id':13,'name':'Project Forge','description':'Robotics & automation upgrade — reduce direct labour by 35% across 4 lines',
     'project_category':'Capital Investment','project_type':'Automation','status':'Approved','owner':'Sandra Cole',
     'budget_type':'CapEx','department':'Manufacturing','region':'Midwest','risk_level':'Medium','strategic_score':4,
     'budget':32.0,'capacity':250.0,'unit_label':'kT','unit_price':9500.0,'unit_cost':1800.0,'efficiency':1.0,
     'ramp':'Fast','schedule_driven':False,'accelerate':False,'locked_year':None,'stages':_stages('Capital Investment')},
    {'id':14,'name':'Project Summit','description':'APAC market entry — establish sales presence in Japan, Korea, Australia',
     'project_category':'Growth / Revenue','project_type':'Market Expansion','status':'Concept','owner':'Claire Dubois',
     'budget_type':'Mixed','department':'Commercial','region':'APAC','risk_level':'High','strategic_score':5,
     'budget':42.0,'capacity':100.0,'unit_label':'k customers','unit_price':900.0,'unit_cost':280.0,'efficiency':0.75,
     'ramp':'Slow','schedule_driven':False,'accelerate':False,'locked_year':None,'stages':_stages('Growth / Revenue')},
    {'id':15,'name':'Project Pulse','description':'AI-powered predictive maintenance — reduce unplanned downtime by 60%',
     'project_category':'Digital / IT','project_type':'AI / ML Platform','status':'Active','owner':'Ravi Patel',
     'budget_type':'OpEx','department':'Maintenance & Reliability','region':'North America','risk_level':'Medium','strategic_score':4,
     'budget':6.5,'capacity':60.0,'unit_label':'machines','unit_price':180000.0,'unit_cost':28000.0,'efficiency':0.85,
     'ramp':'Fast','schedule_driven':False,'accelerate':True,'locked_year':None,'stages':_stages('Digital / IT')},
    {'id':16,'name':'Project Clarity','description':'Carbon neutrality roadmap — renewable energy + carbon offset portfolio',
     'project_category':'ESG / Sustainability','project_type':'Sustainability Program','status':'Approved','owner':'James Okafor',
     'budget_type':'CapEx','department':'EHS & Sustainability','region':'Global','risk_level':'Low','strategic_score':4,
     'budget':22.0,'capacity':80.0,'unit_label':'k tCO2','unit_price':62.0,'unit_cost':9.0,'efficiency':0.90,
     'ramp':'Slow','schedule_driven':True,'accelerate':False,'locked_year':None,'stages':_stages('ESG / Sustainability')},
    {'id':17,'name':'Project Bridge','description':'Post-merger integration — consolidate 6 acquired BUs onto single platform',
     'project_category':'M&A / Integration','project_type':'Integration','status':'Active','owner':'Alex Riordan',
     'budget_type':'CapEx','department':'Corporate Development','region':'North America','risk_level':'High','strategic_score':5,
     'budget':18.0,'capacity':6.0,'unit_label':'BUs','unit_price':5500000.0,'unit_cost':1200000.0,'efficiency':0.85,
     'ramp':'Fast','schedule_driven':True,'accelerate':True,'locked_year':None,'stages':_stages('M&A / Integration')},
    {'id':18,'name':'Project Orbit','description':'R&D Centre of Excellence — dedicated innovation campus, 8 research programs',
     'project_category':'R&D / Innovation','project_type':'Innovation Centre','status':'Concept','owner':'Dr. Lena Fischer',
     'budget_type':'CapEx','department':'R&D','region':'Northeast','risk_level':'High','strategic_score':5,
     'budget':58.0,'capacity':8.0,'unit_label':'programs','unit_price':18000000.0,'unit_cost':5000000.0,'efficiency':0.65,
     'ramp':'Slow','schedule_driven':False,'accelerate':False,'locked_year':None,'stages':_stages('R&D / Innovation')},
]

# ── Data helpers ──────────────────────────────────────────────────────────────

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, encoding='utf-8') as f:
            raw = json.load(f)
        raw['projects'] = [_migrate(p) for p in raw.get('projects', [])]
        # Ensure settings has all new keys
        s = raw.setdefault('settings', {})
        _fill_setting_defaults(s)
        return raw
    data = {
        'projects': copy.deepcopy(DEFAULT_PROJECTS),
        'schedule': {},
        'settings': _default_settings(),
    }
    save_data(data)
    return data

def _default_settings():
    return {
        'start_year':      2025,
        'horizon':         10,
        'annual_budget':   50.0,
        'objective':       'FFCF',
        'custom_weights':  None,
        # Financial model assumptions
        'wacc':            0.10,
        'tax_rate':        0.26,
        'inflation':       0.03,
        'dep_life':        10,
        'currency':        '$',
        # Per-year budget overrides  {str(year): float}
        'per_year_budgets': {},
        # New financial model fields
        'wc_pct':                0.10,
        'cost_inflation':        0.00,
        'terminal_value_ebitda': 0.0,
    }

def _fill_setting_defaults(s):
    defaults = _default_settings()
    for k, v in defaults.items():
        if k not in s:
            s[k] = v

def _migrate(p):
    if 'budget' not in p:
        p['budget'] = p.pop('eac', 0)
    if 'capacity' not in p and 'capacity_kt' in p:
        p['capacity']   = p.pop('capacity_kt', 0)
        p['unit_price'] = p.pop('price', 0) * 1000
        p['unit_cost']  = p.pop('opex_per_ton', 0) * 1000
        p['unit_label'] = 'kT'
    if 'efficiency' not in p:
        p['efficiency'] = p.pop('yield_pct', 1.0)
    if 'project_category' not in p:
        p['project_category'] = 'Capital Investment'
    if 'budget_type' not in p:
        p['budget_type'] = 'CapEx'
    if 'risk_level' not in p:
        p['risk_level'] = 'Medium'
    if 'strategic_score' not in p:
        p['strategic_score'] = 3
    if 'status' not in p:
        p['status'] = 'Concept'
    if 'owner' not in p:
        p['owner'] = ''
    if 'department' not in p:
        p['department'] = ''
    if 'region' not in p:
        p['region'] = ''
    if 'stages' not in p or not p['stages']:
        p['stages'] = copy.deepcopy(
            DEFAULT_STAGES.get(p.get('project_category', 'Capital Investment'),
                               DEFAULT_STAGES['Capital Investment']))
    if 'description' not in p:
        p['description'] = ''
    if 'project_type' not in p:
        p['project_type'] = p.pop('element', 'Other')
    return p

def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)

def next_id(projects):
    return max((p['id'] for p in projects), default=0) + 1

# ── Financial model ───────────────────────────────────────────────────────────

def get_assumptions(settings):
    return {
        'wacc':                  float(settings.get('wacc',      WACC)),
        'tax_rate':              float(settings.get('tax_rate',  TAX_RATE)),
        'inflation':             float(settings.get('inflation', INFLATION)),
        'horizon':               int(  settings.get('horizon',   HORIZON)),
        'dep_life':              int(  settings.get('dep_life',  10)),
        'currency':              str(  settings.get('currency',  '$')),
        'wc_pct':                float(settings.get('wc_pct',                0.10)),
        'cost_inflation':        float(settings.get('cost_inflation',        0.00)),
        'terminal_value_ebitda': float(settings.get('terminal_value_ebitda', 0.0)),
    }

def _irr(fcf, guess=0.20):
    if not any(cf < 0 for cf in fcf) or not any(cf > 0 for cf in fcf):
        return 0.0
    r = guess
    for _ in range(200):
        npv  = sum(cf / (1 + r) ** t for t, cf in enumerate(fcf))
        dnpv = sum(-t * cf / (1 + r) ** (t + 1) for t, cf in enumerate(fcf))
        if abs(dnpv) < 1e-12:
            break
        r_new = r - npv / dnpv
        if abs(r_new - r) < 1e-9:
            r = r_new; break
        r = max(-0.999, min(r_new, 50.0))
    check = sum(cf / (1 + r) ** t for t, cf in enumerate(fcf))
    if abs(check) > 1e6 or r < -0.99:
        return 0.0
    return round(r, 6)

def _normalize(values):
    mn, mx = min(values), max(values)
    return [0.5] * len(values) if mx == mn else [(v - mn) / (mx - mn) for v in values]

def compute_financials(p, asmp=None):
    if asmp is None:
        asmp = get_assumptions({})
    wacc     = asmp['wacc']
    tax      = asmp['tax_rate']
    inf      = asmp['inflation']
    horizon  = asmp['horizon']
    dep_life = asmp['dep_life']
    wc_pct   = asmp['wc_pct']
    cost_inf = asmp['cost_inflation']
    tv_mult  = asmp['terminal_value_ebitda']

    budget_abs = p['budget'] * 1e6
    capacity   = p.get('capacity', 0)
    btype      = p.get('budget_type', 'CapEx')

    if budget_abs == 0 and capacity == 0:
        return {'npv': 0.0, 'irr': 0.0, 'roce': 0.0, 'payback': None,
                'cap_efficiency': 0.0, 'fcf_schedule': [], 'ebit_schedule': []}

    ramp_f = RAMP.get(p.get('ramp', 'Slow'), RAMP['Slow'])
    # Extend ramp if horizon > 10
    if len(ramp_f) < horizon:
        ramp_f = ramp_f + [ramp_f[-1]] * (horizon - len(ramp_f))

    if btype == 'CapEx':
        dep      = budget_abs / dep_life
        year0_cf = -budget_abs
    elif btype == 'OpEx':
        dep      = 0.0
        year0_cf = -budget_abs * (1 - tax)
    else:  # Mixed: 50% capitalised, 50% expensed
        dep      = (budget_abs * 0.5) / dep_life
        year0_cf = -(budget_abs * 0.5) - (budget_abs * 0.5 * (1 - tax))

    fcf, ebit_list, wc = [year0_cf], [], 0.0

    for y in range(1, horizon + 1):
        rf   = ramp_f[min(y - 1, len(ramp_f) - 1)]
        rev  = capacity * p.get('unit_price', 0) * p.get('efficiency', 1.0) * rf * (1 + inf) ** (y - 1)
        cost = capacity * p.get('unit_cost',  0) * rf * (1 + cost_inf) ** (y - 1)
        ebit = rev - cost
        ebit_list.append(ebit)

        taxable = max(0.0, ebit - dep)
        tax_pay = taxable * tax
        cf      = ebit - tax_pay + dep
        if y == 1:
            wc  = wc_pct * rev; cf -= wc
        if y == horizon:
            cf += wc
        fcf.append(cf)

    # Terminal value: add to final year FCF before discounting
    if tv_mult > 0 and ebit_list:
        fcf[-1] += ebit_list[-1] * tv_mult

    npv     = sum(cf / (1 + wacc) ** t for t, cf in enumerate(fcf))
    irr     = _irr(fcf)
    roce    = sum(ebit_list) / budget_abs if budget_abs > 0 else 0.0
    cap_eff = npv / budget_abs if budget_abs > 0 else 0.0

    cumul, payback = 0.0, None
    for t, cf in enumerate(fcf):
        cumul += cf
        if cumul > 0 and payback is None:
            payback = t

    return {
        'npv':            round(npv / 1e6, 3),
        'irr':            round(irr, 4),
        'roce':           round(roce, 4),
        'payback':        payback,
        'cap_efficiency': round(cap_eff, 4),
        'fcf_schedule':   [round(cf / 1e6, 3) for cf in fcf],
        'ebit_schedule':  [round(e  / 1e6, 3) for e  in ebit_list],
    }

# ── Sequencing engines ────────────────────────────────────────────────────────

def _score_and_sort(inv_projs, w):
    if not inv_projs:
        return inv_projs
    npvs    = [p.get('npv',  0) for p in inv_projs]
    irrs    = [p.get('irr',  0) for p in inv_projs]
    roces   = [p.get('roce', 0) for p in inv_projs]
    budgets = [p['budget']      for p in inv_projs]
    n_npv, n_irr, n_roce, n_bud = (
        _normalize(npvs), _normalize(irrs), _normalize(roces), _normalize(budgets))
    for i, p in enumerate(inv_projs):
        s = (w['npv']*n_npv[i] + w['irr']*n_irr[i] +
             w['roce']*n_roce[i] - w['capex']*n_bud[i])
        if p.get('accelerate'):      s += 0.15
        if p.get('schedule_driven'): s += 100.0
        p['score'] = round(s, 4)
    return inv_projs

def _year_cap(year, settings, annual_budget):
    pyr = settings.get('per_year_budgets', {})
    return float(pyr.get(str(year), annual_budget))

def sequence_projects(projects, settings):
    start_year    = int(settings.get('start_year', 2025))
    horizon       = int(settings.get('horizon', 10))
    annual_budget = float(settings.get('annual_budget', 50.0))
    objective     = settings.get('objective', 'FFCF')
    w             = settings.get('custom_weights') or WEIGHTS.get(objective, WEIGHTS['FFCF'])
    end_year      = start_year + horizon
    asmp          = get_assumptions(settings)

    projs = copy.deepcopy(projects)
    for p in projs:
        p.update(compute_financials(p, asmp))

    free_projs = [p for p in projs if p['budget'] == 0 and p.get('capacity', 0) == 0]
    inv_projs  = [p for p in projs if not (p['budget'] == 0 and p.get('capacity', 0) == 0)]

    _score_and_sort(inv_projs, w)

    year_data = {y: {'budget_used': 0.0, 'projects': [],
                     'cap': _year_cap(y, settings, annual_budget)}
                 for y in range(start_year, end_year + 1)}
    schedule  = {}

    for p in inv_projs:
        ly = p.get('locked_year')
        if ly and start_year <= ly < end_year:
            schedule[str(p['id'])] = ly
            year_data[ly]['budget_used'] += p['budget']
            year_data[ly]['projects'].append(p['id'])

    to_place = sorted([p for p in inv_projs if str(p['id']) not in schedule],
                      key=lambda p: -p.get('score', 0))

    for year in range(start_year, end_year):
        cap  = year_data[year]['cap']
        left = cap - year_data[year]['budget_used']
        next_q = []
        for p in to_place:
            if p['budget'] <= left + 1e-9:
                schedule[str(p['id'])] = year
                year_data[year]['budget_used'] += p['budget']
                year_data[year]['projects'].append(p['id'])
                left -= p['budget']
            else:
                next_q.append(p)
        to_place = next_q

    for p in free_projs:
        ly = p.get('locked_year') or start_year
        ly = max(start_year, min(ly, end_year - 1))
        schedule[str(p['id'])] = ly
        year_data[ly]['projects'].append(p['id'])

    return _build_result(projs, schedule, to_place, year_data)

def apply_manual_schedule(projects, settings):
    assignments   = {str(k): v for k, v in settings.get('manual_assignments', {}).items()}
    start_year    = int(settings.get('start_year', 2025))
    horizon       = int(settings.get('horizon', 10))
    annual_budget = float(settings.get('annual_budget', 50.0))
    end_year      = start_year + horizon
    asmp          = get_assumptions(settings)

    projs = copy.deepcopy(projects)
    for p in projs:
        p.update(compute_financials(p, asmp))
        p['score'] = None

    year_data = {y: {'budget_used': 0.0, 'projects': [],
                     'cap': _year_cap(y, settings, annual_budget)}
                 for y in range(start_year, end_year + 1)}
    schedule  = {}

    for p in projs:
        yr = assignments.get(str(p['id']))
        if yr and start_year <= int(yr) < end_year:
            yr = int(yr)
            schedule[str(p['id'])] = yr
            year_data[yr]['budget_used'] += p['budget']
            year_data[yr]['projects'].append(p['id'])

    unplaced = [p for p in projs if str(p['id']) not in schedule]
    return _build_result(projs, schedule, unplaced, year_data)

def _build_result(projs, schedule, unplaced, year_data):
    unscheduled_ids = [p['id'] for p in unplaced]
    for p in projs:
        p['scheduled_year'] = schedule.get(str(p['id']))
        p['unscheduled']    = p['id'] in unscheduled_ids

    all_fcf_by_year = {}
    for p in projs:
        if p['scheduled_year'] and p.get('fcf_schedule'):
            base = p['scheduled_year']
            for offset, cf in enumerate(p['fcf_schedule']):
                yr = base + offset
                all_fcf_by_year[yr] = all_fcf_by_year.get(yr, 0.0) + cf

    return projs, year_data, unscheduled_ids, all_fcf_by_year

# ── API helpers ───────────────────────────────────────────────────────────────

def _build_opt_response(result_projects, year_data, unscheduled_ids, portfolio_fcf, settings):
    proj_map = {p['id']: p for p in result_projects}

    years = sorted(year_data.keys())
    year_summary = []
    for y in years:
        yd = year_data[y]
        plist = [proj_map[pid] for pid in yd['projects'] if pid in proj_map]
        year_summary.append({
            'year':          y,
            'budget_used':   round(yd['budget_used'], 2),
            'budget_cap':    round(yd.get('cap', settings.get('annual_budget', 50.0)), 2),
            'project_ids':   yd['projects'],
            'project_names': [p['name'] for p in plist],
        })

    sorted_yrs = sorted(portfolio_fcf)
    cumul, cumul_fcf = 0.0, []
    for yr in sorted_yrs:
        cumul += portfolio_fcf[yr]
        cumul_fcf.append({'year': yr, 'cumulative_fcf': round(cumul, 2)})

    breakeven = next((c['year'] for c in cumul_fcf if c['cumulative_fcf'] > 0), None)
    return {
        'projects':       result_projects,
        'year_summary':   year_summary,
        'cumulative_fcf': cumul_fcf,
        'unscheduled':    unscheduled_ids,
        'breakeven_year': breakeven,
        'settings':       settings,
    }

# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/projects', methods=['GET'])
def get_projects():
    data = load_data()
    asmp = get_assumptions(data['settings'])
    projs = copy.deepcopy(data['projects'])
    for p in projs:
        p.update(compute_financials(p, asmp))
    return jsonify(projs)

@app.route('/api/projects', methods=['POST'])
def add_project():
    data = load_data()
    asmp = get_assumptions(data['settings'])
    p = _migrate(request.json)
    p['id'] = next_id(data['projects'])
    data['projects'].append(p)
    save_data(data)
    p.update(compute_financials(p, asmp))
    return jsonify(p), 201

@app.route('/api/projects/<int:pid>', methods=['PUT'])
def update_project(pid):
    data = load_data()
    asmp = get_assumptions(data['settings'])
    for i, p in enumerate(data['projects']):
        if p['id'] == pid:
            updated = _migrate(request.json)
            updated['id'] = pid
            data['projects'][i] = updated
            save_data(data)
            updated.update(compute_financials(updated, asmp))
            return jsonify(updated)
    return jsonify({'error': 'not found'}), 404

@app.route('/api/projects/<int:pid>', methods=['DELETE'])
def delete_project(pid):
    data = load_data()
    data['projects'] = [p for p in data['projects'] if p['id'] != pid]
    save_data(data)
    return jsonify({'ok': True})

@app.route('/api/settings', methods=['GET', 'POST'])
def settings_route():
    data = load_data()
    if request.method == 'POST':
        data['settings'].update(request.json)
        save_data(data)
    return jsonify(data['settings'])

@app.route('/api/stages/defaults', methods=['GET'])
def get_default_stages():
    return jsonify(DEFAULT_STAGES)

@app.route('/api/optimize', methods=['POST'])
def optimize():
    body     = request.json or {}
    data     = load_data()
    settings = {**data['settings'], **body}

    if settings.get('objective') == 'Manual' and settings.get('manual_assignments'):
        rp, yd, un, pf = apply_manual_schedule(data['projects'], settings)
    else:
        rp, yd, un, pf = sequence_projects(data['projects'], settings)

    resp = _build_opt_response(rp, yd, un, pf, settings)

    data['schedule'] = {str(p['id']): p['scheduled_year'] for p in rp}
    data['settings'] = {k: v for k, v in settings.items() if k != 'manual_assignments'}
    save_data(data)
    return jsonify(resp)

@app.route('/api/metrics', methods=['GET'])
def metrics():
    data  = load_data()
    asmp  = get_assumptions(data['settings'])
    projs = copy.deepcopy(data['projects'])
    for p in projs:
        p.update(compute_financials(p, asmp))

    inv = [p for p in projs if p['budget'] > 0 or p.get('capacity', 0) > 0]
    total_npv  = round(sum(p.get('npv',  0) for p in inv), 2)
    total_bud  = round(sum(p['budget']       for p in inv), 2)
    irrs  = [p.get('irr',  0) for p in inv if p.get('irr',  0) > 0]
    roces = [p.get('roce', 0) for p in inv if p.get('roce', 0) > 0]
    avg_irr  = round(sum(irrs)  / len(irrs),  4) if irrs  else 0
    avg_roce = round(sum(roces) / len(roces), 4) if roces else 0

    ranked = sorted(inv, key=lambda p: -(p.get('cap_efficiency', 0)))
    for i, p in enumerate(ranked):
        p['rank'] = i + 1

    cat_counts = {}
    status_counts = {}
    for p in projs:
        c = p.get('project_category', 'Other')
        cat_counts[c] = cat_counts.get(c, 0) + 1
        s = p.get('status', 'Concept')
        status_counts[s] = status_counts.get(s, 0) + 1

    # Health alerts
    health_alerts = []
    for p in inv:
        irr_v = p.get('irr', 0)
        if irr_v > 0 and irr_v < asmp['wacc']:
            health_alerts.append({'level': 'warn', 'project': p['name'],
                'msg': f"IRR {irr_v*100:.1f}% is below WACC {asmp['wacc']*100:.1f}% — may destroy value"})
        if p.get('npv', 0) < 0 and p.get('risk_level') in ('High', 'Critical'):
            health_alerts.append({'level': 'danger', 'project': p['name'],
                'msg': f"Negative NPV ({p['npv']:.1f}M) with {p['risk_level']} risk"})
        if not p.get('owner', '').strip():
            health_alerts.append({'level': 'info', 'project': p['name'], 'msg': 'No owner assigned'})

    # Category budget breakdown
    cat_budget = {}
    for p in projs:
        c = p.get('project_category', 'Other')
        if c not in cat_budget:
            cat_budget[c] = {'count': 0, 'budget': 0.0, 'npv': 0.0}
        cat_budget[c]['count'] += 1
        cat_budget[c]['budget'] = round(cat_budget[c]['budget'] + p['budget'], 2)
        cat_budget[c]['npv']    = round(cat_budget[c]['npv']    + p.get('npv', 0), 2)

    return jsonify({
        'total_npv':      total_npv,
        'total_budget':   total_bud,
        'avg_irr':        avg_irr,
        'avg_roce':       avg_roce,
        'project_count':  len(projs),
        'category_counts':cat_counts,
        'status_counts':  status_counts,
        'ranked':         ranked,
        'assumptions':    asmp,
        'settings':       data['settings'],
        'health_alerts':  health_alerts,
        'cat_budget':     cat_budget,
    })

@app.route('/api/sensitivity', methods=['GET'])
def sensitivity():
    data  = load_data()
    base  = get_assumptions(data['settings'])
    projs = data['projects']

    # WACC levels: base ± up to 6pp
    base_w = base['wacc']
    levels = [max(0.01, base_w + d) for d in [-0.06, -0.04, -0.02, 0, 0.02, 0.04, 0.06]]

    portfolio = []
    for w in levels:
        a2  = {**base, 'wacc': w}
        npv = sum(compute_financials(p, a2)['npv'] for p in projs)
        portfolio.append({'wacc': round(w, 4), 'total_npv': round(npv, 2)})

    # Per-project sensitivity for top-8 by budget
    top8 = sorted(projs, key=lambda p: -p.get('budget', 0))[:8]
    projects_sens = []
    for p in top8:
        row = {'name': p['name'], 'base_npv': compute_financials(p, base)['npv']}
        for w in levels:
            row[f'w{int(round(w*100))}'] = compute_financials(p, {**base, 'wacc': w})['npv']
        projects_sens.append(row)

    return jsonify({
        'portfolio':    portfolio,
        'projects':     projects_sens,
        'base_wacc':    base_w,
        'wacc_levels':  [round(w, 4) for w in levels],
    })

@app.route('/api/export/csv')
def export_csv():
    data  = load_data()
    asmp  = get_assumptions(data['settings'])
    cur   = asmp['currency']
    si    = io.StringIO()
    w     = csv.writer(si)
    w.writerow(['Name','Owner','Status','Category','Type','Budget Type','Department','Region',
                f'Budget ({cur}M)','Risk','Strategic Score',
                f'NPV ({cur}M)','IRR (%)','ROCE (x)','Payback (yr)','Cap Efficiency',
                'Ramp','Accelerate','Schedule Driven','Locked Year'])
    for p in data['projects']:
        fin = compute_financials(p, asmp)
        w.writerow([
            p['name'], p.get('owner',''), p.get('status',''), p['project_category'],
            p.get('project_type',''), p['budget_type'], p.get('department',''), p.get('region',''),
            p['budget'], p['risk_level'], p.get('strategic_score',''),
            fin['npv'], round(fin['irr']*100, 2), round(fin['roce'], 4),
            fin['payback'], round(fin['cap_efficiency'], 4),
            p.get('ramp',''), p.get('accelerate',''), p.get('schedule_driven',''), p.get('locked_year',''),
        ])
    return Response(
        si.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment;filename=caplogix_portfolio.csv'}  # filename kept lowercase for compatibility
    )

@app.route('/api/export/xlsx')
def export_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import datetime
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500

    data  = load_data()
    asmp  = get_assumptions(data['settings'])
    cur   = asmp['currency']
    projs = copy.deepcopy(data['projects'])
    for p in projs:
        p.update(compute_financials(p, asmp))
    inv   = [p for p in projs if p['budget'] > 0 or p.get('capacity', 0) > 0]

    wb = Workbook()

    # ── helpers ──────────────────────────────────────────────────────────────
    def hdr(cell, text, bg='1E40AF'):
        cell.value = text
        cell.font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        cell.fill  = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def val(cell, v, bold=False, color=None, bg=None, num_fmt=None, align='left'):
        cell.value = v
        cell.font  = Font(name='Calibri', size=10, bold=bold,
                          color=color if color else '000000')
        if bg:  cell.fill = PatternFill('solid', fgColor=bg)
        if num_fmt: cell.number_format = num_fmt
        cell.alignment = Alignment(horizontal=align)

    thin = Side(style='thin', color='E2E8F0')
    def border(cell):
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── SHEET 1: Executive Summary ────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Executive Summary'
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells('A1:F1')
    ws['A1'].value = 'CapLogiX — Strategic Portfolio Report'
    ws['A1'].font  = Font(name='Calibri', size=20, bold=True, color='1E3A5F')
    ws['A1'].fill  = PatternFill('solid', fgColor='EFF6FF')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 40

    ws['A2'].value = f'Generated: {datetime.datetime.now().strftime("%d %b %Y  %H:%M")}'
    ws['A2'].font  = Font(name='Calibri', size=9, color='64748B', italic=True)

    # ── KPIs
    total_npv  = round(sum(p.get('npv',  0) for p in inv), 1)
    total_bud  = round(sum(p['budget']       for p in inv), 1)
    irrs       = [p.get('irr', 0) for p in inv if p.get('irr', 0) > 0]
    avg_irr    = round(sum(irrs)/len(irrs)*100, 1) if irrs else 0
    roces      = [p.get('roce',0) for p in inv if p.get('roce',0) > 0]
    avg_roce   = round(sum(roces)/len(roces), 2) if roces else 0

    kpis = [
        ('Total Portfolio NPV', f'{cur}{total_npv:,.1f}M', '16a34a' if total_npv >= 0 else 'dc2626'),
        ('Total Budget',        f'{cur}{total_bud:,.1f}M',  '1E40AF'),
        ('Avg IRR',             f'{avg_irr:.1f}%',           '0d9488'),
        ('Avg ROCE',            f'{avg_roce:.2f}x',          '7c3aed'),
        ('# Projects',          str(len(projs)),              '374151'),
        ('# Categories',        str(len({p["project_category"] for p in projs})), '374151'),
    ]
    for col_i, (label, value, color) in enumerate(kpis):
        c = col_i + 1
        ws.cell(row=4, column=c, value=label).font  = Font(name='Calibri', size=9, color='64748B', bold=True)
        ws.cell(row=4, column=c).alignment = Alignment(horizontal='center')
        vc = ws.cell(row=5, column=c, value=value)
        vc.font  = Font(name='Calibri', size=16, bold=True, color=color)
        vc.fill  = PatternFill('solid', fgColor='F8FAFC')
        vc.alignment = Alignment(horizontal='center')
        ws.row_dimensions[5].height = 30
        ws.column_dimensions[get_column_letter(c)].width = 20

    # ── Financial Assumptions
    ws['A7'].value = 'FINANCIAL MODEL ASSUMPTIONS'
    ws['A7'].font  = Font(name='Calibri', bold=True, size=10, color='1E3A5F')
    asmp_items = [
        ('WACC',              f"{asmp['wacc']*100:.1f}%"),
        ('Corporate Tax Rate',f"{asmp['tax_rate']*100:.1f}%"),
        ('Revenue Inflation', f"{asmp['inflation']*100:.1f}%"),
        ('Cost Inflation',    f"{asmp.get('cost_inflation',0)*100:.1f}%"),
        ('Dep. Life (CapEx)', f"{asmp['dep_life']} yrs"),
        ('Working Capital',   f"{asmp.get('wc_pct',0.10)*100:.0f}% of Y1 Rev"),
        ('Terminal Value',    f"{asmp.get('terminal_value_ebitda',0):.1f}x EBITDA" if asmp.get('terminal_value_ebitda',0) > 0 else 'None'),
        ('Horizon',           f"{asmp['horizon']} years"),
        ('Currency',          f"{cur}M"),
    ]
    for i, (label, value) in enumerate(asmp_items):
        r = 8 + i
        ws.cell(row=r, column=1, value=label).font  = Font(name='Calibri', size=10, color='374151')
        ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='F8FAFC')
        ws.cell(row=r, column=2, value=value).font  = Font(name='Calibri', size=10, bold=True)

    # ── Status breakdown
    ws.cell(row=7, column=4, value='STATUS BREAKDOWN').font = Font(name='Calibri', bold=True, size=10, color='1E3A5F')
    status_cnt = {}
    for p in projs:
        s = p.get('status', 'Concept')
        status_cnt[s] = status_cnt.get(s, 0) + 1
    for i, (s, cnt) in enumerate(sorted(status_cnt.items())):
        r = 8 + i
        ws.cell(row=r, column=4, value=s).font = Font(name='Calibri', size=10)
        ws.cell(row=r, column=5, value=cnt).font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='center')

    # ── Category budget table
    start_r = 7 + max(len(asmp_items), len(status_cnt)) + 2
    ws.cell(row=start_r, column=1, value='BUDGET & NPV BY CATEGORY').font = Font(name='Calibri', bold=True, size=10, color='1E3A5F')
    for col_i, label in enumerate(['Category','# Projects',f'Budget ({cur}M)',f'NPV ({cur}M)'], 1):
        hdr(ws.cell(row=start_r+1, column=col_i), label)
    cat_grp = {}
    for p in projs:
        c = p.get('project_category','Other')
        cat_grp.setdefault(c, {'count':0,'budget':0.0,'npv':0.0})
        cat_grp[c]['count'] += 1
        cat_grp[c]['budget'] += p['budget']
        cat_grp[c]['npv'] += p.get('npv', 0)
    for i, (cat, d) in enumerate(sorted(cat_grp.items(), key=lambda x: -x[1]['budget'])):
        r = start_r + 2 + i
        bg = 'F8FAFC' if i % 2 == 0 else 'FFFFFF'
        val(ws.cell(r,1), cat, bg=bg)
        val(ws.cell(r,2), d['count'], align='center', bg=bg)
        val(ws.cell(r,3), round(d['budget'],1), num_fmt='#,##0.0', align='right', bg=bg)
        npv_bg = 'D1FAE5' if d['npv'] >= 0 else 'FEE2E2'
        val(ws.cell(r,4), round(d['npv'],1), num_fmt='#,##0.0', align='right',
            color='065F46' if d['npv'] >= 0 else '991B1B', bg=npv_bg)

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    # ── SHEET 2: Project Portfolio ────────────────────────────────────────────
    ws2 = wb.create_sheet('Project Portfolio')
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'A2'

    hdrs2 = ['Project Name','Owner','Status','Category','Type','Budget Type','Dept',
             f'Budget ({cur}M)','Risk','Strategic Score',
             f'NPV ({cur}M)','IRR (%)','ROCE (x)','Payback (yr)','Cap Eff','Ramp']
    ws2.row_dimensions[1].height = 28
    for c, h in enumerate(hdrs2, 1):
        hdr(ws2.cell(1, c), h)
        ws2.column_dimensions[get_column_letter(c)].width = [24,14,12,20,16,12,16,12,10,8,12,10,10,10,10,8][c-1]

    status_bg = {'Active':'D1FAE5','Approved':'DBEAFE','Concept':'F1F5F9',
                 'On Hold':'FEF3C7','Completed':'CCFBF1','Cancelled':'FEE2E2'}
    risk_bg   = {'Low':'D1FAE5','Medium':'FEF3C7','High':'FEE2E2','Critical':'FEE2E2'}

    for r, p in enumerate(sorted(projs, key=lambda x: -(x.get('npv',0))), 2):
        alt = 'F8FAFC' if r % 2 == 0 else 'FFFFFF'
        data_vals = [
            p['name'], p.get('owner',''), p.get('status','Concept'), p.get('project_category',''),
            p.get('project_type',''), p['budget_type'], p.get('department',''),
            p['budget'], p.get('risk_level',''), p.get('strategic_score',3),
            p.get('npv',0), round(p.get('irr',0)*100,2), round(p.get('roce',0),3),
            p.get('payback',''), round(p.get('cap_efficiency',0),3), p.get('ramp',''),
        ]
        for c, v in enumerate(data_vals, 1):
            cell = ws2.cell(r, c, v)
            cell.font = Font(name='Calibri', size=10,
                             bold=(c == 1),
                             color='991B1B' if c == 11 and isinstance(v, (int,float)) and v < 0 else
                                   '065F46' if c == 11 and isinstance(v, (int,float)) and v >= 0 else '000000')
            cell.fill = PatternFill('solid', fgColor=(
                status_bg.get(v, alt) if c == 3 else
                risk_bg.get(v, alt)   if c == 9 else
                ('D1FAE5' if isinstance(v,(int,float)) and v >= 0 else 'FEE2E2') if c == 11 else alt))
            cell.alignment = Alignment(horizontal='center' if c in (2,3,6,9,10,14,16) else 'right' if c in (8,11,12,13,15) else 'left')
            if c in (8,11,12,13): cell.number_format = '#,##0.00'

    # ── SHEET 3: FCF Schedules ─────────────────────────────────────────────────
    ws3 = wb.create_sheet('FCF Schedules')
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = 'C2'

    hz = asmp['horizon']
    ws3.row_dimensions[1].height = 28
    hdr(ws3.cell(1,1), 'Project')
    hdr(ws3.cell(1,2), 'Type')
    for t in range(hz+1):
        hdr(ws3.cell(1, t+3), f'Y{t}' if t < hz else f'Y{t}*')
    hdr(ws3.cell(1, hz+4), f'NPV ({cur}M)')
    hdr(ws3.cell(1, hz+5), 'IRR (%)')
    hdr(ws3.cell(1, hz+6), 'ROCE (x)')

    ws3.column_dimensions['A'].width = 26
    ws3.column_dimensions['B'].width = 8
    for t in range(hz+4):
        ws3.column_dimensions[get_column_letter(t+3)].width = 10

    portfolio_fcf = [0.0] * (hz+1)
    for r, p in enumerate(sorted(projs, key=lambda x: x.get('project_category','')), 2):
        alt = 'F8FAFC' if r % 2 == 0 else 'FFFFFF'
        ws3.cell(r,1,p['name']).font  = Font(name='Calibri', size=10, bold=True)
        ws3.cell(r,1).fill            = PatternFill('solid', fgColor=alt)
        ws3.cell(r,2,p['budget_type']).font = Font(name='Calibri', size=9, color='64748B')
        ws3.cell(r,2).fill            = PatternFill('solid', fgColor=alt)
        fcf_sch = p.get('fcf_schedule', [])
        for t in range(hz+1):
            cf = fcf_sch[t] if t < len(fcf_sch) else 0.0
            if t < len(portfolio_fcf): portfolio_fcf[t] = round(portfolio_fcf[t] + cf, 2)
            c = ws3.cell(r, t+3, round(cf,2))
            c.number_format = '#,##0.00'
            c.font  = Font(name='Calibri', size=10,
                           color='DC2626' if cf < 0 else '16A34A' if cf > 0 else '64748B')
            c.fill  = PatternFill('solid', fgColor=alt)
            c.alignment = Alignment(horizontal='right')
        for col_i, v in enumerate([round(p.get('npv',0),2), round(p.get('irr',0)*100,2), round(p.get('roce',0),3)], hz+4):
            c2 = ws3.cell(r, col_i, v)
            c2.number_format = '#,##0.00'
            c2.font = Font(name='Calibri', size=10,
                           color=('065F46' if v >= 0 else '991B1B') if col_i == hz+4 else '000000')
            c2.fill = PatternFill('solid', fgColor=('D1FAE5' if v >= 0 else 'FEE2E2') if col_i == hz+4 else alt)
            c2.alignment = Alignment(horizontal='right')

    # Portfolio total row
    tr = len(projs) + 2
    hdr(ws3.cell(tr,1), 'PORTFOLIO TOTAL', bg='1E40AF')
    hdr(ws3.cell(tr,2), '',               bg='1E40AF')
    cumul = 0.0
    for t, cf in enumerate(portfolio_fcf):
        cumul += cf
        c = ws3.cell(tr, t+3, round(cf, 2))
        c.number_format = '#,##0.00'
        c.font  = Font(name='Calibri', bold=True, color='FFFFFF')
        c.fill  = PatternFill('solid', fgColor='1E40AF')
        c.alignment = Alignment(horizontal='right')
    # Cumulative row
    cr = tr + 1
    ws3.cell(cr,1,'Cumulative FCF').font = Font(name='Calibri', bold=True, color='1E40AF')
    ws3.cell(cr,1).fill = PatternFill('solid', fgColor='EFF6FF')
    cumul2 = 0.0
    for t, cf in enumerate(portfolio_fcf):
        cumul2 += cf
        c = ws3.cell(cr, t+3, round(cumul2, 2))
        c.number_format = '#,##0.00'
        c.font  = Font(name='Calibri', size=10, bold=True,
                       color='065F46' if cumul2 >= 0 else 'DC2626')
        c.fill  = PatternFill('solid', fgColor='EFF6FF')
        c.alignment = Alignment(horizontal='right')

    ws3.cell(hz+4, hz+4, f'* Y{hz} includes terminal value if set').font = Font(name='Calibri', size=8, color='94A3B8', italic=True)

    # ── SHEET 4: Sensitivity Analysis ─────────────────────────────────────────
    ws4 = wb.create_sheet('Sensitivity Analysis')
    ws4.sheet_view.showGridLines = False

    ws4['A1'].value = 'WACC Sensitivity Analysis'
    ws4['A1'].font  = Font(name='Calibri', bold=True, size=14, color='1E40AF')
    ws4['A2'].value = f'Base WACC: {asmp["wacc"]*100:.1f}%   |   Shows total portfolio NPV at alternative WACC assumptions'
    ws4['A2'].font  = Font(name='Calibri', size=10, color='64748B', italic=True)

    base_w = asmp['wacc']
    wacc_levels = [max(0.01, base_w + d) for d in [-0.06,-0.04,-0.02, 0, 0.02, 0.04, 0.06]]

    # Portfolio row header
    hdr(ws4.cell(4,1), 'Metric', bg='1E3A5F')
    for ci, w in enumerate(wacc_levels, 2):
        lbl = f'{w*100:.0f}%' + (' (base)' if abs(w - base_w) < 0.001 else '')
        hdr(ws4.cell(4, ci), lbl, bg='1E40AF' if abs(w - base_w) < 0.001 else '374151')

    # Portfolio NPV row
    ws4.cell(5,1,'Total Portfolio NPV').font = Font(name='Calibri', bold=True, size=10)
    ws4.cell(5,1).fill = PatternFill('solid', fgColor='F8FAFC')
    for ci, w in enumerate(wacc_levels, 2):
        a2  = {**asmp, 'wacc': w}
        npv = round(sum(compute_financials(p, a2)['npv'] for p in data['projects']), 1)
        c = ws4.cell(5, ci, npv)
        c.number_format = f'"{cur}"#,##0.0'
        c.font  = Font(name='Calibri', size=10, bold=abs(w-base_w) < 0.001,
                       color='065F46' if npv >= 0 else 'DC2626')
        c.fill  = PatternFill('solid', fgColor='D1FAE5' if npv >= 0 else 'FEE2E2')
        c.alignment = Alignment(horizontal='center')

    # Per-project sensitivity
    ws4['A7'].value = 'Per-Project NPV by WACC Level'
    ws4['A7'].font  = Font(name='Calibri', bold=True, size=11, color='1E3A5F')
    hdr(ws4.cell(8,1), 'Project')
    hdr(ws4.cell(8,2), f'Base NPV ({cur}M)')
    for ci, w in enumerate(wacc_levels, 3):
        hdr(ws4.cell(8, ci), f'{w*100:.0f}%')

    top8 = sorted(data['projects'], key=lambda p: -p.get('budget',0))[:8]
    for ri, p in enumerate(top8, 9):
        alt = 'F8FAFC' if ri % 2 == 0 else 'FFFFFF'
        ws4.cell(ri,1,p['name']).font = Font(name='Calibri', size=10, bold=True)
        ws4.cell(ri,1).fill = PatternFill('solid', fgColor=alt)
        base_npv = compute_financials(p, asmp)['npv']
        c = ws4.cell(ri,2,round(base_npv,2))
        c.number_format = '#,##0.00'
        c.font  = Font(name='Calibri', size=10, bold=True,
                       color='065F46' if base_npv >= 0 else 'DC2626')
        c.fill  = PatternFill('solid', fgColor='D1FAE5' if base_npv >= 0 else 'FEE2E2')
        c.alignment = Alignment(horizontal='right')
        for ci, w in enumerate(wacc_levels, 3):
            npv_w = round(compute_financials(p, {**asmp, 'wacc': w})['npv'], 2)
            c2 = ws4.cell(ri, ci, npv_w)
            c2.number_format = '#,##0.00'
            c2.font = Font(name='Calibri', size=10,
                           color='065F46' if npv_w >= 0 else 'DC2626')
            c2.fill = PatternFill('solid', fgColor=alt)
            c2.alignment = Alignment(horizontal='right')

    ws4.column_dimensions['A'].width = 28
    for ci in range(2, len(wacc_levels)+4):
        ws4.column_dimensions[get_column_letter(ci)].width = 14

    # ── Output ────────────────────────────────────────────────────────────────
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(
        bio.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=caplogix_portfolio.xlsx'}
    )

# ── Scenarios ─────────────────────────────────────────────────────────────────

def _safe_sc_name(name):
    """Sanitise scenario name → safe filename (alphanumeric + space/dash/underscore)."""
    return re.sub(r'[^\w \-]', '', name).strip()[:80]

def _sc_path(safe_name):
    os.makedirs(SCENARIOS_DIR, exist_ok=True)
    return os.path.join(SCENARIOS_DIR, safe_name + '.json')

@app.route('/api/scenarios', methods=['GET'])
def list_scenarios():
    os.makedirs(SCENARIOS_DIR, exist_ok=True)
    out = []
    for fname in sorted(os.listdir(SCENARIOS_DIR)):
        if not fname.endswith('.json'):
            continue
        fpath = os.path.join(SCENARIOS_DIR, fname)
        try:
            with open(fpath, encoding='utf-8') as f:
                sc = json.load(f)
            projs = sc.get('projects', [])
            out.append({
                'name':          sc.get('name', fname[:-5]),
                'saved_at':      sc.get('saved_at', ''),
                'project_count': len(projs),
                'total_budget':  round(sum(p.get('budget', 0) for p in projs), 1),
                'total_npv':     round(sum(p.get('npv', 0) for p in projs), 1),
                'objective':     sc.get('settings', {}).get('objective', '—'),
            })
        except Exception:
            pass
    return jsonify(out)

@app.route('/api/scenarios/save', methods=['POST'])
def save_scenario():
    body = request.get_json() or {}
    raw_name = body.get('name', '').strip()
    if not raw_name:
        return jsonify({'error': 'Name is required'}), 400
    safe = _safe_sc_name(raw_name)
    if not safe:
        return jsonify({'error': 'Invalid name — use letters, numbers, spaces or dashes'}), 400
    data = load_data()
    sc = {
        'name':     raw_name,
        'saved_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
        'projects': data['projects'],
        'settings': data['settings'],
    }
    with open(_sc_path(safe), 'w', encoding='utf-8') as f:
        json.dump(sc, f, indent=2)
    return jsonify({'ok': True, 'name': raw_name})

@app.route('/api/scenarios/load/<path:name>', methods=['POST'])
def load_scenario(name):
    safe = _safe_sc_name(name)
    fpath = _sc_path(safe)
    if not os.path.exists(fpath):
        return jsonify({'error': 'Scenario not found'}), 404
    with open(fpath, encoding='utf-8') as f:
        sc = json.load(f)
    data = load_data()
    data['projects'] = [_migrate(p) for p in sc.get('projects', [])]
    if sc.get('settings'):
        data['settings'] = sc['settings']
        _fill_setting_defaults(data['settings'])
    save_data(data)
    return jsonify({'ok': True, 'name': sc.get('name', name)})

@app.route('/api/scenarios/delete/<path:name>', methods=['DELETE'])
def delete_scenario(name):
    safe = _safe_sc_name(name)
    fpath = _sc_path(safe)
    if os.path.exists(fpath):
        os.remove(fpath)
    return jsonify({'ok': True})

@app.route('/api/reset', methods=['POST'])
def reset_data():
    if os.path.exists(DATA_FILE):
        os.remove(DATA_FILE)
    load_data()
    return jsonify({'ok': True})

# ── Launch ────────────────────────────────────────────────────────────────────

# Render (and other cloud hosts) inject PORT; absence means we're running locally.
_PORT        = int(os.environ.get('PORT', 5000))
_IS_LOCAL    = 'PORT' not in os.environ

def _open_browser():
    webbrowser.open(f'http://127.0.0.1:{_PORT}')

if __name__ == '__main__':
    if _IS_LOCAL:
        threading.Timer(1.2, _open_browser).start()
        print('\n  +-----------------------------------------+')
        print(f'  |  CapLogiX  ->  http://127.0.0.1:{_PORT}       |')
        print('  +-----------------------------------------+\n')
    app.run(debug=False, host='0.0.0.0', port=_PORT)
